"""
FISCAM 2024 Internal Control Test Plan generator.

Reproduces the DLA/GAO-style two-tab workbook used for audit-quality control
testing (matches the user's Enterprise_SM_04_02_01.xlsx sample):

  Tab 1  "Instructions - Test Plan"  -- purpose, roles, conventions, authoritative docs
  Tab 2  "Test Plan"                 -- 44 numbered line items in three sections:
         Risks and Internal Control Details (1-18)
         Test of Design / TOD          (19-25)
         Test of Effectiveness / TOE   (26-44)

Division of labor (the standing guardrail):
  * The LLM drafts PROSE/auto fields from the FISCAM control (control objective,
    identified risks, control activity description, TOD/TOE procedures, etc.).
  * Code owns the structure, line numbering, and section layout.
  * Human owns signatures, dates, Pass/Fail results, and rationale -- left blank.
  * Metadata that doesn't apply to IT controls defaults to "Not Applicable".

Output is a draft for Preparer/Tester completion -- never an attestation.
"""

from __future__ import annotations

import datetime as _dt
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .llm_provider import LLMProvider, get_provider

# --------------------------------------------------------------------------- #
# Template: (line_number | None for section header, label, field_key | None)
# --------------------------------------------------------------------------- #
SECTION = object()  # marker

TEMPLATE: list[tuple[Any, str, str | None]] = [
    (SECTION, "Risks and Internal Control Details", None),
    (1,  "J/D Code / MSC", "jd_code"),
    (2,  "Assessable Unit", "assessable_unit"),
    (3,  "Sub-Assessable Units", "sub_assessable_units"),
    (4,  "Assessable Unit Manager", "assessable_unit_manager"),
    (5,  "Control ID #", "control_id"),
    (6,  "Operation Objective/Financial Objective", "operation_objective"),
    (7,  "Identified Risks", "identified_risks"),
    (8,  "Risk Categories (Strategic, Financial, Regulatory and Compliance, Reputational, Operational, Fraud, Cyber, Data, Improper Payment)", "risk_categories"),
    (9,  "Control Objective", "control_objective"),
    (10, "Management Assertion (for Financial Objective)", "management_assertion"),
    (11, "Impacted Financial Statement Line Item (for Financial Objective)", "impacted_fsli"),
    (12, "Control Type (Manual, Automated, Combined, Preventive, Detective)", "control_type"),
    (13, "Control Category (Management Review, Segregation of Duties, Reconciliations, Authorizations/Approvals, Systems Edit Check/Validations)", "control_category"),
    (14, "Control Activity Description", "control_activity_description"),
    (15, "Control Frequency", "control_frequency"),
    (16, "Location of Executed Control", "location_executed"),
    (17, "Preparer", "preparer"),
    (18, "Date", "preparer_date"),
    (SECTION, "Test of Design (TOD)", None),
    (19, "Testing Date", "tod_testing_date"),
    (20, "Internal Control Document Used for TOD Procedures", "tod_documents"),
    (21, "Walkthrough Locations", "tod_walkthrough_locations"),
    (22, "Test Procedures", "tod_procedures"),
    (23, "Test Results (Pass/Fail)", "tod_result"),
    (24, "Rational for TOD Pass/Fail", "tod_rationale"),
    (25, "Tester's Name and Signature", "tod_tester"),
    (SECTION, "Test of Effectiveness (TOE)", None),
    (SECTION, "Note: TOE section is completed ONLY if TOD passes per line 23", None),
    (26, "Testing Date", "toe_testing_date"),
    (27, "Test Method (Inquiry, Observation, Inspection/Re-Performance)", "toe_test_method"),
    (28, "Testing Period", "toe_testing_period"),
    (29, "Testing Location", "toe_testing_location"),
    (30, "Population Description", "toe_population"),
    (31, "Population Completeness Validation", "toe_population_completeness"),
    (32, "Sample Size", "toe_sample_size"),
    (33, "Number of Acceptable Deviations", "toe_acceptable_deviations"),
    (34, "Tools Used to Select Random Sample", "toe_sample_tools"),
    (35, "Source Documents Used for Testing", "toe_source_documents"),
    (36, "Test Procedures", "toe_procedures"),
    (37, "Test Results (Pass/Fail)", "toe_result"),
    (38, "Rational for TOE Pass/Fail", "toe_rationale"),
    (39, "Compensating Control", "compensating_control"),
    (40, "Tester's Name and Signature", "toe_tester"),
    (41, "Validator's Name", "validator_name"),
    (42, "Date of Validation", "validation_date"),
    (43, "RICA's Name and Signature", "rica_name"),
    (44, "RICA's Approval Date", "rica_date"),
]

# Fields a human completes during execution -- left blank in the draft.
_HUMAN_FIELDS = {
    "assessable_unit_manager", "preparer", "preparer_date",
    "tod_testing_date", "tod_result", "tod_rationale", "tod_tester",
    "toe_testing_date", "toe_testing_period", "toe_result", "toe_rationale",
    "toe_tester", "validator_name", "validation_date", "rica_name", "rica_date",
}
# Fields the LLM drafts from the control.
_LLM_FIELDS = {
    "identified_risks", "risk_categories", "control_objective", "control_type",
    "control_category", "control_activity_description", "control_frequency",
    "tod_documents", "tod_procedures", "toe_test_method", "toe_population",
    "toe_population_completeness", "toe_sample_size", "toe_acceptable_deviations",
    "toe_sample_tools", "toe_source_documents", "toe_procedures",
}
# List-valued fields rendered as "1. ...\n2. ..."
_LIST_FIELDS = {"tod_documents", "tod_procedures", "toe_source_documents"}

# Org-context defaults (override via org_context dict).
# Blank by default — client/engagement fields are filled when a client is known
# (via org_context or directly in the workbook). No org-identifying defaults.
_ORG_DEFAULTS = {
    "jd_code": "",
    "assessable_unit": "",
    "sub_assessable_units": "",
    "operation_objective": "",
    "management_assertion": "",
    "impacted_fsli": "",
    "location_executed": "",
    "tod_walkthrough_locations": "",
    "toe_testing_location": "",
    "compensating_control": "",
}

SYSTEM_PROMPT = """\
You are a FISCAM 2024 internal-control test-plan preparer. Given a FISCAM 2024
control, draft the design and testing fields for an audit-quality test plan.

Return ONLY valid JSON with these keys (strings unless noted):
- identified_risks: the risk(s) the control mitigates
- risk_categories: from {Strategic, Financial, Regulatory and Compliance,
  Reputational, Operational, Fraud, Cyber, Data, Improper Payment}
- control_objective: the FISCAM control objective
- control_type: from {Manual, Automated, Combined, Preventive, Detective}
- control_category: from {Management Review, Segregation of Duties,
  Reconciliations, Authorizations/Approvals, Systems Edit Check/Validations}
- control_activity_description: concrete description of how the control operates
- control_frequency: e.g. "Annually (and upon significant system change)"
- tod_documents: LIST of internal control documents to inspect for design testing
- tod_procedures: LIST of Test-of-Design steps (inquire/inspect/compare)
- toe_test_method: e.g. "Inquiry, Inspection, Re-Performance"
- toe_population: description of the testing population
- toe_population_completeness: how completeness of the population is validated
- toe_sample_size: e.g. "100% of in-scope systems"
- toe_acceptable_deviations: typically "0"
- toe_sample_tools: tool/method for sample selection
- toe_source_documents: LIST of source documents used for effectiveness testing
- toe_procedures: Test-of-Effectiveness procedure description

Be specific and verifiable. Do not attest to compliance.
"""


@dataclass
class FiscamTestPlan:
    control_id: str
    fields: dict[str, str] = field(default_factory=dict)
    provenance: dict[str, Any] = field(default_factory=dict)
    needs_review: bool = True


def _render_list(val: Any) -> str:
    if isinstance(val, list):
        return "\n".join(f"{i+1}. {str(x).strip()}" for i, x in enumerate(val))
    return str(val)


def _fiscam_control(conn, control_id: str) -> dict | None:
    """Pull the FISCAM control text if it's loaded (current version)."""
    if conn is None:
        return None
    row = conn.execute(
        """SELECT c.title, c.statement, c.guidance FROM controls c
             JOIN catalog_versions v ON v.catalog_version_id=c.catalog_version_id
            WHERE v.framework_id='fiscam' AND c.control_id=? AND v.is_current=1""",
        (control_id.lower(),)).fetchone()
    return dict(row) if row else None


def _nist_800_53a_overrides(conn, control_id: str) -> dict[str, str]:
    """If control_id is an 800-53 control, build authoritative TOD/TOE fields from
    its embedded 800-53A objectives + Examine/Interview/Test methods. Returns {} if
    the control has no 800-53A content (e.g. a FISCAM control)."""
    if conn is None:
        return {}
    try:
        from . import assessment
    except Exception:
        return {}
    objs = assessment.objectives(conn, control_id)
    if not objs:
        return {}
    meths = {m.method: m for m in assessment.methods(conn, control_id)}
    examine = meths.get("EXAMINE")
    test = meths.get("TEST")
    interview = meths.get("INTERVIEW")

    tod = ["Inquire with responsible personnel to confirm the control is designed "
           "to meet each 800-53A assessment objective below."]
    tod += [f"Determine that {o.prose}" for o in objs]

    docs = (examine.objects if examine else []) or [
        "Policy, procedures, and other relevant documents for this control."]
    method_label = ", ".join(
        lbl for lbl, present in (("Inquiry/Interview", interview),
                                 ("Inspection/Examine", examine),
                                 ("Re-Performance/Test", test)) if present
    ) or "Inquiry, Inspection"
    toe_steps = []
    if test:
        toe_steps.append("Test (re-perform): " + "; ".join(test.objects[:6]))
    if interview:
        toe_steps.append("Interview: " + "; ".join(interview.objects[:6]))
    toe_steps.append("For each sampled item, gather current-period evidence showing "
                     "each assessment objective is met; document exceptions.")

    return {
        "control_objective": (objs[0].prose if len(objs) == 1
                              else f"Satisfy all {len(objs)} NIST SP 800-53A "
                                   f"assessment objectives for this control."),
        "tod_procedures": _render_list(tod),
        "tod_documents": _render_list(docs),
        "toe_test_method": method_label,
        "toe_source_documents": _render_list(docs),
        "toe_procedures": "\n".join(toe_steps),
    }


def draft_test_plan(
    conn, *, control_id: str,
    control_title: str = "", control_text: str = "",
    org_context: dict[str, str] | None = None,
    provider: LLMProvider | None = None,
    use_800_53a: bool = True,
) -> FiscamTestPlan:
    """Draft a FISCAM 2024 test plan for one control. needs_review always.

    If use_800_53a and the control is an 800-53 control, the TOD/TOE procedures,
    documents, and test methods are taken from the authoritative NIST 800-53A
    assessment objectives instead of LLM prose."""
    provider = provider or get_provider()
    org = {**_ORG_DEFAULTS, **(org_context or {})}

    overrides = _nist_800_53a_overrides(conn, control_id) if use_800_53a else {}

    ctrl = _fiscam_control(conn, control_id) or {}
    title = control_title or ctrl.get("title", "")
    text = control_text or ctrl.get("statement", "")

    prompt = (f"FISCAM 2024 Control {control_id} -- {title}\n"
              f"Control text: {text}\n"
              f"Guidance: {ctrl.get('guidance','')}\n\n"
              "Draft the test-plan fields as JSON.")
    result = provider.complete(system=SYSTEM_PROMPT, prompt=prompt, effort="high")

    parsed = _parse_json(result.text)

    det = _deterministic_fields(control_id, title, text, org)

    fields: dict[str, str] = {}
    # org / metadata
    for k, v in org.items():
        fields[k] = v
    fields["control_id"] = control_id
    # LLM-drafted -> use structured LLM output where present, else deterministic draft
    for k in _LLM_FIELDS:
        if k in parsed and parsed[k] not in (None, "", []):
            fields[k] = _render_list(parsed[k]) if k in _LIST_FIELDS else str(parsed[k])
        else:
            fields.setdefault(k, det[k])
    # authoritative NIST 800-53A overrides win over LLM/deterministic prose
    fields.update(overrides)
    # human fields -> blank for completion
    for k in _HUMAN_FIELDS:
        fields.setdefault(k, "")

    now = _dt.datetime.now(_dt.timezone.utc).isoformat()
    if overrides:
        source = "nist_800_53a" + ("+llm" if parsed else "")
    else:
        source = "llm" if parsed else "deterministic_template"
    return FiscamTestPlan(
        control_id=control_id, fields=fields, needs_review=True,
        provenance={**result.provenance(), "drafted_at": now,
                    "structured": bool(parsed), "field_source": source,
                    "authoritative_800_53a": bool(overrides)})


def _deterministic_fields(control_id: str, title: str, text: str,
                          org: dict[str, str]) -> dict[str, str]:
    """Full FISCAM-style draft without an LLM. TOD/TOE boilerplate is standard;
    control-specific fields derive from the control text. Always needs_review."""
    unit = org.get("assessable_unit") or "the assessable unit"
    objective = title or (text[:120] if text else "the control objective")
    activity = text or "[provide the control activity description]"
    std_docs = [
        "System Security Plan (SSP) — current approved version (eMASS)",
        "Authoritative policy/SOP governing the control",
        "Approval memorandum / management sign-off",
        "Risk Assessment Report (RAR) from eMASS",
    ]
    return {
        "identified_risks": f"Risk that {objective.lower()} is not achieved, "
            "weakening the control environment and exposing the organization to "
            "regulatory, cyber, and operational impact.",
        "risk_categories": "Regulatory and Compliance; Cyber",
        "control_objective": objective,
        "control_type": "Manual / Preventive",
        "control_category": "Authorizations/Approvals",
        "control_activity_description": activity,
        "control_frequency": "Annually (and upon significant system change)",
        "tod_documents": _render_list(std_docs),
        "tod_procedures": _render_list([
            "Inquire with responsible personnel to understand how the control operates.",
            "Inspect the governing policy/SOP and supporting documentation.",
            "Compare the documented control against the FISCAM 2024 control objective.",
            "Confirm management review/approval of the control's results.",
            "Verify the control was performed within its stated frequency.",
        ]),
        "toe_test_method": "Inquiry, Inspection, Re-Performance",
        "toe_population": f"In-scope systems within {unit} subject to this control.",
        "toe_population_completeness": "Reconcile the population to the source-of-record "
            "(eMASS, ACAS Reporting Center, or applicable system inventory).",
        "toe_sample_size": "100% of in-scope systems (typically a small population — "
            "sample size = N).",
        "toe_acceptable_deviations": "0",
        "toe_sample_tools": "Not applicable — full population tested.",
        "toe_source_documents": _render_list(std_docs),
        "toe_procedures": "For each sampled item, re-perform the TOD procedures against "
            "current-period evidence; document results and note any exceptions.",
    }


def _parse_json(text: str) -> dict[str, Any]:
    text = text.strip()
    s, e = text.find("{"), text.rfind("}")
    if s == -1 or e == -1:
        return {}
    try:
        return json.loads(text[s:e + 1])
    except json.JSONDecodeError:
        return {}


# --------------------------------------------------------------------------- #
# Workbook writer (openpyxl) -- reproduces the two-tab template
# --------------------------------------------------------------------------- #
def write_workbook(plan: FiscamTestPlan, path: str | Path,
                   authoritative_docs: list[str] | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    path = Path(path)
    title_fill = PatternFill("solid", fgColor="E8E1AA")
    sect_fill = PatternFill("solid", fgColor="E2EFDA")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    wrap_top_left = Alignment(wrap_text=True, vertical="top", horizontal="left")

    wb = Workbook()

    # ---- Tab 1: Instructions ----
    ins = wb.active
    ins.title = "Instructions - Test Plan"
    ins.column_dimensions["A"].width = 30
    ins.column_dimensions["B"].width = 110
    docs = authoritative_docs or _DEFAULT_AUTH_DOCS
    rows = [
        (f"Internal Control Test Plan — Instructions ({plan.control_id})", None),
        ("Purpose", "This workbook documents the design and operating effectiveness testing of the FISCAM 2024 control on the Test Plan tab, structured to support audit-quality evidence collection aligned to GAO's revised FISCAM (2024) framework."),
        ("Workbook Structure", "Tab 1 (this tab): Instructions and definitions.\nTab 2 (Test Plan): Pre-populated control details, plus Test of Design (TOD) and Test of Effectiveness (TOE) sections to be completed by the Preparer / Tester."),
        ("Roles", "Preparer: Documents control design and TOD evidence. Marks TOD Pass/Fail and signs.\nTester: Performs TOE sample testing if TOD passes; documents results, signs.\nValidator: Independent review of test procedures and conclusions.\nRICA: Reviewing Internal Control Assessor — final approval signature."),
        ("Completion Sequence", "1. Preparer reviews/updates Risks and Internal Control Details (rows 1-18).\n2. Preparer completes Test of Design (rows 19-25). If TOD = Pass, proceed to TOE.\n3. Tester completes Test of Effectiveness (rows 26-44).\n4. Validator reviews and signs (rows 41-42).\n5. RICA approves and signs (rows 43-44)."),
        ("Key Conventions", "• \"Not Applicable\" in row 3 and row 6 means those metadata fields don't apply — it does NOT mean the control itself is NA.\n• If the Operation/Financial Objective is NA, rows 10 and 11 are also NA.\n• TOE is completed ONLY if TOD passes per row 23.\n• Compensating Control (row 39) is completed ONLY when TOE fails."),
        ("FISCAM 2024 Framework Note", "This template implements FISCAM 2024 control language and illustrative procedures. Prior-period testing performed under FISCAM 2009 is not directly transferable; current-period testing must be performed against FISCAM 2024 illustrative controls."),
        ("Authoritative Documents (for TOD inspection)", "\n".join(f"• {d}" for d in docs)),
    ]
    for i, (a, b) in enumerate(rows, start=1):
        ca = ins.cell(i, 1, a); ca.alignment = wrap_top_left; ca.font = bold
        if b is not None:
            cb = ins.cell(i, 2, b); cb.alignment = wrap
        else:
            ca.fill = title_fill

    # ---- Tab 2: Test Plan ----
    ws = wb.create_sheet("Test Plan")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 90

    r = 1
    t = ws.cell(r, 1, "Internal Control Test Plan"); t.font = bold; t.fill = title_fill
    t.alignment = wrap
    r += 1
    for row in TEMPLATE:
        marker, label, key = row
        if marker is SECTION:
            c = ws.cell(r, 1, label); c.font = bold; c.fill = sect_fill; c.alignment = wrap
        else:
            num = marker
            ws.cell(r, 1, num).alignment = wrap
            ws.cell(r, 2, label).alignment = wrap
            ws.cell(r, 3, plan.fields.get(key, "")).alignment = wrap
        r += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


_DEFAULT_AUTH_DOCS = [
    "Organizational RMF / cybersecurity policy and standard operating procedures",
    "DoD Manual 5200.01, Information Security Program (if applicable)",
    "FIPS 199 / CNSSI 1253 (security categorization standards)",
    "System Security Plan (SSP) and Risk Assessment Report (RAR)",
    "eMASS — Enterprise Mission Assurance Support Service",
    "ACAS — Assured Compliance Assessment Solution",
]
