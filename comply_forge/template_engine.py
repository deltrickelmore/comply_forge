"""
Bring-your-own-template engine for FISCAM test plans (the "upload a template" button).

Instead of regenerating a workbook from our built-in layout, the user uploads
THEIR workbook; we learn where each field lives and fill values into the existing
cells -- preserving their exact formatting, dropdowns, branding, and extra tabs.

How it works:
  learn_template(path) inspects the sheet and builds a TemplateSpec mapping each
  ComplyForge field_key -> the cell (row, col) to write into. Matching is anchored
  on the line-number column (1..N) when present -- robust against repeated labels
  like "Test Procedures"/"Date" across TOD and TOE. Falls back to section-aware
  label matching for unnumbered templates.

  fill_template(src, plan_fields, out) clones the uploaded file and writes the
  drafted values into the mapped cells (skipping human sign-off fields and any
  field with no value), then saves. Formatting and other tabs are preserved.
"""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .fiscam_test_plan import TEMPLATE, SECTION, _HUMAN_FIELDS

# number -> canonical field_key (from the built-in template definition)
_NUMBER_TO_KEY: dict[int, str] = {
    row[0]: row[2] for row in TEMPLATE if row[0] is not SECTION and row[2]
}

# Distinctive label keywords -> field_key, for unnumbered templates. Order matters
# (more specific first). Section-qualified keys are resolved at match time.
_KEYWORD_TO_KEY: list[tuple[str, str]] = [
    ("sub-assessable", "sub_assessable_units"),
    ("assessable unit manager", "assessable_unit_manager"),
    ("assessable unit", "assessable_unit"),
    ("j/d code", "jd_code"), ("jd code", "jd_code"),
    ("control id", "control_id"),
    ("operation objective", "operation_objective"),
    ("identified risk", "identified_risks"),
    ("risk categor", "risk_categories"),
    ("control objective", "control_objective"),
    ("management assertion", "management_assertion"),
    ("impacted financial", "impacted_fsli"), ("fsli", "impacted_fsli"),
    ("control type", "control_type"),
    ("control categor", "control_category"),
    ("control activity", "control_activity_description"),
    ("control frequency", "control_frequency"),
    ("location of executed", "location_executed"),
    ("internal control document", "tod_documents"),
    ("walkthrough", "tod_walkthrough_locations"),
    ("test method", "toe_test_method"),
    ("testing period", "toe_testing_period"),
    ("testing location", "toe_testing_location"),
    ("population description", "toe_population"),
    ("population completeness", "toe_population_completeness"),
    ("sample size", "toe_sample_size"),
    ("acceptable deviations", "toe_acceptable_deviations"),
    ("tools used to select", "toe_sample_tools"),
    ("source documents", "toe_source_documents"),
    ("compensating control", "compensating_control"),
    ("validator", "validator_name"),
    ("rica", "rica_name"),
    ("preparer", "preparer"),
]
# Labels that depend on TOD vs TOE section context.
_SECTION_KEYWORDS = {
    "test procedures": {"tod": "tod_procedures", "toe": "toe_procedures"},
    "test results": {"tod": "tod_result", "toe": "toe_result"},
    "rational": {"tod": "tod_rationale", "toe": "toe_rationale"},
    "testing date": {"tod": "tod_testing_date", "toe": "toe_testing_date"},
    "tester's name": {"tod": "tod_tester", "toe": "toe_tester"},
    "date": {"tod": "preparer_date", "toe": None},  # bare "Date" -> preparer date
}


def _norm(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


@dataclass
class TemplateSpec:
    source: str
    sheet: str
    label_col: int
    value_col: int
    number_col: int | None
    field_rows: dict[str, int] = field(default_factory=dict)   # field_key -> row
    unmatched: list[tuple[int, str]] = field(default_factory=list)  # (row, label)

    def summary(self) -> dict[str, Any]:
        return {"source": self.source, "sheet": self.sheet,
                "value_col": self.value_col, "fields_mapped": len(self.field_rows),
                "unmatched_labels": len(self.unmatched)}


def _pick_sheet(wb, sheet: str | None):
    if sheet:
        return wb[sheet]
    # prefer a sheet that looks like a test plan (has a numbering column or many labels)
    best, best_score = wb.worksheets[0], -1
    for ws in wb.worksheets:
        labels = sum(1 for r in range(1, min(ws.max_row, 60) + 1)
                     for c in range(1, min(ws.max_column, 6) + 1)
                     if isinstance(ws.cell(r, c).value, str) and len(str(ws.cell(r, c).value)) > 3)
        if labels > best_score:
            best, best_score = ws, labels
    return best


def _find_number_col(ws) -> int | None:
    """Column whose cells form a run of integers 1,2,3,... (the line-number col)."""
    for c in range(1, min(ws.max_column, 6) + 1):
        seq = [ws.cell(r, c).value for r in range(1, min(ws.max_row, 60) + 1)]
        ints = [v for v in seq if isinstance(v, int)]
        if len(ints) >= 10 and ints[:3] == [1, 2, 3]:
            return c
    return None


def learn_template(path: str | Path, sheet: str | None = None) -> TemplateSpec:
    from openpyxl import load_workbook
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = _pick_sheet(wb, sheet)

    num_col = _find_number_col(ws)
    if num_col is not None:
        label_col, value_col = num_col + 1, num_col + 2
    else:
        label_col, value_col = 1, 2  # assume label|value in first two columns

    spec = TemplateSpec(source=str(path), sheet=ws.title,
                        label_col=label_col, value_col=value_col, number_col=num_col)

    section = "tod"  # default; updated as we walk section headers
    for r in range(1, ws.max_row + 1):
        num = ws.cell(r, num_col).value if num_col else None
        label = ws.cell(r, label_col).value
        nlabel = _norm(label)

        # track TOD/TOE section for unnumbered / section-dependent labels
        head = _norm(ws.cell(r, 1).value)
        if "test of design" in head:
            section = "tod"
        elif "test of effectiveness" in head:
            section = "toe"

        if num_col and isinstance(num, int) and num in _NUMBER_TO_KEY:
            spec.field_rows[_NUMBER_TO_KEY[num]] = r
            continue
        if not nlabel:
            continue

        key = _match_label(nlabel, section)
        if key:
            spec.field_rows.setdefault(key, r)
        elif label and num_col is None:  # only log misses for label-based templates
            spec.unmatched.append((r, str(label)))

    return spec


def _match_label(nlabel: str, section: str) -> str | None:
    for kw, mapping in _SECTION_KEYWORDS.items():
        if kw in nlabel:
            return mapping.get(section)
    for kw, key in _KEYWORD_TO_KEY:
        if kw in nlabel:
            return key
    return None


def fill_template(src: str | Path, plan_fields: dict[str, str], out: str | Path,
                  spec: TemplateSpec | None = None,
                  include_human_fields: bool = False) -> Path:
    """Clone the uploaded template and write drafted values into mapped cells.
    Human sign-off fields are left as-is unless include_human_fields=True."""
    from openpyxl import load_workbook
    src, out = Path(src), Path(out)
    spec = spec or learn_template(src)

    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(src, out)                # preserve everything verbatim first
    wb = load_workbook(out)                  # keep formulas/styles
    ws = wb[spec.sheet]

    written = 0
    for key, row in spec.field_rows.items():
        if key in _HUMAN_FIELDS and not include_human_fields:
            continue
        val = plan_fields.get(key)
        if val in (None, ""):
            continue
        ws.cell(row, spec.value_col, val)    # style of the existing cell is retained
        written += 1
    wb.save(str(out))
    return out
